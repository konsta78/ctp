from django.shortcuts import render, redirect, reverse
from django.views import View
from django.views import generic
from .models import Employee, AdressDepartment, Department, SubdivisionDepartament
from django.contrib.auth.models import User, Group
from openpyxl import load_workbook
from django.template.defaulttags import register
from . import addresses
import re


# регистрация нового фильтра для шаблона.
# передача в шаблон значения из словаря по ключу
@register.filter(name="get_item")
def get_item(dictionary, key):
    return dictionary.get(key)


class IndexView(View):
    """
    Отображение стартовой страницы
    """

    def get(self, request):


        # num_visits = request.session.get('num_visits', 0)
        # request.session['num_visits'] = num_visits + 1

        employees = Employee.objects.all()
        departments = Department.objects.all()
        sub_departments = SubdivisionDepartament.objects.all()
        context = {"employees": employees,
                   "departments": departments,
                   "sub_departments": sub_departments}
        return render(request, 'catalog/index.html', context)


class GovernanceView(View):
    """
    Отображение страницы с информацией о руководстве
    """
    def get(self, request):
        employees = Employee.objects.all()
        departments = Department.objects.filter(name='Руководство')
        context = {"employees": employees, "departments": departments}
        return render(request, 'catalog/index.html', context)


class DepartmentsView(View):
    """
    Отображение страницы с информацией об отделах
    """
    def get(self, request):
        departments = Department.objects.all().order_by('name')
        context = {"departments": departments}
        return render(request, 'catalog/departments.html', context)


class AddressesView(View):
    """
    Отображение странцицы с информацией об отделах, сгруппированных по адресам
    """
    def get(self, request):
        count_dep = {}  # словарь с кол-вом отделов по каждому адресу
        departments = Department.objects.all().order_by('name')
        for i in range(AdressDepartment.objects.count()):
            count = len(Department.objects.filter(adress=AdressDepartment.objects.get(id=i + 1)))
            count_dep[i + 1] = count

        adresses = AdressDepartment.objects.all().order_by('id')
        context = {"addresses": adresses, "departments": departments, "count_dep": count_dep}
        return render(request, 'catalog/addresses.html', context)


class ResultsView(View):
    """
    Отображение страницы с результатами поиска по базе данных
    """
    def check_find(self, request):
        """
        Функция определения контекста и выбора шаблона для отображения
        :param request: запрос пользователя
        :return: context, template
        """
        employees = Employee.objects.all()
        departments = Department.objects.all()

        emp_changes_name = {}  # словарь с измененными ФИО сотрудников (подстветка строки поиска)
        dep_changes_name = {}  # словарь с измененными наименованиями отделов (подстветка строки поиска)

        search = request.GET.get('search')

        if search == "":  # при пустом поле поиска возращаемся на главную страницу
            context = {"employees": employees, "departments": departments}
            template = 'catalog/index.html'
        else:
            objects_employess = []  # список сотрудников по результатам поиска
            objects_departments = []  # список отделов по результатам поиска

            for employee in employees:  # поиск среди сотрудников по ФИО
                if search.lower() in str(employee).lower():
                    tmp = str(employee).split()
                    for i in range(len(tmp)):
                        if tmp[i].lower().startswith(search.lower()):
                            tmp[i] = tmp[i].lower().replace(search.lower(),
                            "<span style='background-color: yellow'>" + search.capitalize() + "</span>")
                        else:
                            tmp[i] = tmp[i].replace(search.lower(),
                            "<span style='background-color: yellow'>" + search.lower() + "</span>").capitalize()
                    tmp_str = " ".join(tmp)
                    emp_changes_name[employee.id] = tmp_str
                    objects_employess.append(employee)

            for department in departments:  # поиск отделов по их наименованию
                if search.lower() in department.name.lower():
                    if department.name.lower().startswith(search.lower()):
                        tmp_str = department.name.lower().replace(search.lower(),
                            "<span style='background-color: yellow'>" + search.capitalize() + "</span>")
                    else:
                        tmp_str = department.name.lower().replace(search.lower(),
                            "<span style='background-color: yellow'>" + search.lower() + "</span>").capitalize()

                    dep_changes_name[department.id] = tmp_str  # наименование отдела с подстветкой строки поиска
                    objects_departments.append(department)

            context = {"employees": objects_employess, "departments": objects_departments,
                       "count_emp": len(objects_employess), "count_dep": len(objects_departments),
                       "dep_changes_name": dep_changes_name, "emp_changes_name": emp_changes_name}
            template = 'catalog/results.html'

        return context, template

    def get(self, request):
        context, template = self.check_find(request)
        return render(request, template, context)


class EmployeeDetail(generic.DetailView):
    """
    Отображение страницы с подробной информацией о сотруднике
    """
    model = Employee


class DepartmentDetail(generic.DetailView):
    """
    Отображение страницы с подробной информацией об отделе
    """
    model = Department

    def get_context_data(self, **kwargs):
        """
        Функция для добавления в контекст об отделе также и данных о сотрудниках
        (обеспечивает работоспособность ссылок на странице 'departments-detail.html')
        :param kwargs:
        :return: context
        """
        context = super().get_context_data(**kwargs)
        context['employees'] = Employee.objects.all()
        context['sub_departments'] = SubdivisionDepartament.objects.all()
        return context


class LoadDataBaseView(View):
    """
    Загрузка данных из excel-файла и отображение на текущей странице
    """
    def read_excel_file(self, file):
        """
        Функция открытия/чтения excel-файла и создания списков ячеек
        :param file: имя входного excel-файла
        :return: рабочую страницу и списки horizontal_merged, departments_cells
        """
        start_cell = 4  # начальная ячейка для парсинга
        wb = load_workbook(file)
        sheet = wb.active
        vertical_merged = []  # список номеров ячеек. являющихся первыми в вертикальных объединенных группах
        horizontal_merged = []  # список номеров горизонтальных объединенных ячеек и их значений
        departments_cells = []  # в список ячеек добавляется размер объединенной группы

        # создание списков vertical_merged и horizontal_merged
        for cell in range(start_cell, sheet.max_row + 1):
            value = sheet[f'A{cell}'].value
            if value is not None:
                vertical_merged.append(cell)
            if not str(sheet[f'B{cell}'].value).isdigit():
                horizontal_merged.append([cell - start_cell + 1, sheet[f'B{cell}'].value])

        # добавляем в список номер последней строки для корректного подсчета диапазона
        vertical_merged.append(sheet.max_row)

        # добавляем размер объединенной группы
        for i in range(1, len(vertical_merged)):
            departments_cells.append([vertical_merged[i - 1],
                                      vertical_merged[i] - vertical_merged[i - 1]])

        return sheet, horizontal_merged, departments_cells

    def create_addresses(self):
        """
        Функция создание записей в БД с адресами объектов
        :return: запись данных в БД
        """
        street_lst = addresses.street_lst
        buildeing_lst = addresses.buildeing_lst
        letter_lst = addresses.letter_lst
        room_lst = addresses.room_lst

        for i in range(len(street_lst)):
            AdressDepartment.objects.update_or_create(
                id=i + 1,
                street=street_lst[i],
                building=buildeing_lst[i],
                letter=letter_lst[i],
                room=room_lst[i]
            )

    def create_subdivision_departament(self, horizontal_merged):
        """
        Функция создание записей в БД с подотделами
        :param horizontal_merged: список горизонтальных объединенных ячеек с наименованиями подотделов
        :return: запись данных в БД
        """
        for k in range(len(horizontal_merged)):
            SubdivisionDepartament.objects.update_or_create(
                id=k + 1,
                sub_name=horizontal_merged[k][1],
                id_employee=horizontal_merged[k][0] + 1
            )

    def create_departments_and_employees(self, sheet, departments_cells):
        """
        Функция создания записей отделов и сотрудников в базе данных
        :param sheet: рабочий лист excel-файла
        :param departments_cells: список с объединенными ячейками с наименованиями отделов
        :return: запись данных в БД
        """
        def create_departments():
            """
            Создание отделов
            :return: запись данных в БД
            """
            Department.objects.update_or_create(
                id=index + 1,
                name=sheet[f'A{departments_cells[index][0]}'].value,
                adress=AdressDepartment.objects.get(id=address_id),
                priority=index
            )

        def create_employees():
            """
            Создание сотрудников
            :return: запись данных в БД
            """
            for i in range(departments_cells[index][1]):
                current_cell = departments_cells[index][0] + i

                # обработка пустых ячеек столбца 'С'
                if sheet[f'C{current_cell}'].value is None:
                    pass
                    # print('Поле "С" пустое!')

                # обработка пустых ячеек столбца 'D'
                elif sheet[f'D{current_cell}'].value is None:
                    pass
                    # fio = [" ", " ", " "]
                else:
                    fio = str(sheet[f'D{current_cell}'].value).split()  # список с ФИО
                    if len(fio) < 3:
                        fio.append(' ')
                    Employee.objects.update_or_create(
                        id=current_cell - 3,
                        surname=fio[0],
                        name=fio[1],
                        patronymic=fio[2],
                        position=sheet[f'C{current_cell}'].value,
                        department=Department.objects.get(id=index + 1),
                        phone_work=sheet[f'E{current_cell}'].value,
                        phone_work_additional=sheet[f'F{current_cell}'].value,
                        phone_mob=sheet[f'G{current_cell}'].value,
                        email=sheet[f'H{current_cell}'].value,
                        office=sheet[f'I{current_cell}'].value,
                        dob=None
                    )

        # поиск адресов в ячейках excel-файла и назначение их соответствующим отделам
        for index in range(len(departments_cells)):
            address_id = 1
            if sheet[f'J{departments_cells[index][0]}'].value is not None:
                tmp = sheet[f'J{departments_cells[index][0]}'].value
                nums = "д. " + re.findall(r'\d+', tmp)[0]  # поиск по номеру дома
                for k in range(len(addresses.buildeing_lst)):
                    if nums == addresses.buildeing_lst[k]:
                        address_id = k + 1

            create_departments()
            create_employees()

    def get(self, request):
        """
        Отображение страницы с результатами загрузки данных из excel-файла
        :param request: запрос пользователя
        :return: render / redirect
        """
        if request.user.is_superuser:
            self.create_addresses()
            sheet, horizontal_merged, departments_cells = self.read_excel_file('test.xlsx')
            self.create_subdivision_departament(horizontal_merged)
            self.create_departments_and_employees(sheet, departments_cells)

            employees = Employee.objects.all()
            departments = Department.objects.all()
            sub_departments = SubdivisionDepartament.objects.all()
            context = {"employees": employees,
                       "departments": departments,
                       "sub_departments": sub_departments}
            return render(request, 'catalog/index.html', context)

        else:
            return redirect(reverse('home'))


class DeleteDataBaseView(View):
    """
    Отображение страницы при удалении БД
    """
    def get(self, request):
        if request.user.is_superuser:
            Employee.objects.all().delete()
            SubdivisionDepartament.objects.all().delete()
            Department.objects.all().delete()
            AdressDepartment.objects.all().delete()
        return redirect(reverse('home'))


class UsersListView(View):
    """
    Создание аккаунтов сотрудников, имеющих рабочий e-mail
    и добавление их в группу "Employees"
    """
    def get(self, request):
        if request.user.is_superuser:
            employees = Employee.objects.all()
            group = Group.objects.get(name="Employees")

            for employee in employees:
                if employee.email:
                    username = employee.email
                    try:
                        user = User.objects.get(username=username)
                        if not user:
                            user = User.objects.create_user(username, employee.email, '0000')
                            user.first_name = employee.surname
                            user.last_name = employee.name
                            if employee.patronymic:
                                user.last_name += " " + employee.patronymic
                            user.save()
                            group.user_set.add(user)
                            print("created ", employee)
                    except:
                        pass
        return redirect(reverse('home'))


class UsersDeleteView(View):
    """
    Удаление всех аккаунтов сотрудников кроме аккаунта суперпользователя
    """
    def get(self, request):
        if request.user.is_superuser:
            User.objects.filter(is_superuser=False).delete()
        return redirect(reverse('home'))
